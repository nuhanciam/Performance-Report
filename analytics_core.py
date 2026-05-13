import json
import time
from io import BytesIO
from collections import defaultdict

import pandas as pd
import requests

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


API_VERSION = "2024-04"
SHOP_TIMEZONE = "Europe/Paris"
REQUEST_DELAY_SECONDS = 0.60
MAX_RETRIES = 6


REPORT_ROWS = [
    ("Sessions", "sessions"),
    ("Visteurs uniques", "unique_visitors"),
    ("Sessions / visitor", "sessions_per_visitor"),
    ("Duration (s)", "duration_seconds"),
    ("Duration (min)", "duration_minutes"),
    ("Bounce %", "bounce_pct"),
    ("", None),
    ("Add to cart", "add_to_cart"),
    ("Checkout", "checkout"),
    ("COMMANDES", "# Commandes"),
    ("Conversion %", "conversion_pct"),
    ("Commandes Nx clients", "# Nouveaux clients"),
    ("% cdes Nx clients", "% cdes Nx clients"),
    ("", None),
    ("# Clients uniques", "# Clients"),
    ("# Nx clients", "# Nouveaux clients"),
    ("", None),
    ("# Clients récurrents", "# Clients récurrents"),
    ("", None),
    ("# Produits vendus", "# Produits vendus"),
    ("Ratio cdes/clients", "Ratio cdes/clients"),
    ("Ratio produit / cde", "Ratio produit/cde"),
    ("Ratio produit / client", "Ratio produit/client"),
    ("", None),
    ("Gross sales", "Gross Sales (€)"),
    ("Discounts", "Discounts (€)"),
    ("Discounts %", "Discounts %"),
    ("", None),
    ("Net Sales", "Net Sales (€)"),
    ("Shipping", "Shipping (€)"),
    ("Taxes", "Taxes (€)"),
    ("TOTAL SALES", "Total Sales (€)"),
    ("POIDS %", "POIDS %"),
    ("", None),
    ("AOV HT", "AOV HT (€)"),
    ("AOV TTC (incl ship)", "AOV TTC incl. ship (€)"),
    ("", None),
    ("CA / client HT", "CA/client HT (€)"),
    ("CA / client TTC (incl ship)", "CA/client TTC (€)"),
    ("CA Nx clients HT", "CA Nvx clients HT (€)"),
    ("% CA Nvx clients", "% CA Nvx clients"),
    ("", None),
    ("Frequence achat", "Fréquence achat"),
    ("LTV estimée", "LTV estimée (€)"),
    ("", None),
    ("Retours €", "Retours (€)"),
    ("Retours #", "Retours (#)"),
    ("Retours %", "Retours %"),
    ("", None),
    ("NET SALES", "Net Sales après retours (€)"),
    ("FRANCE", "France (€)"),
    ("", None),
    ("EXPORT", "Export (€)"),
    ("", None),
    ("TOP 3 EXPORT", None),
    ("#1", "Export #1"),
    ("#2", "Export #2"),
    ("#3", "Export #3"),
    ("", None),
    ("TOP 5 PDCT", None),
    ("#1", "Top Pdct #1"),
    ("#2", "Top Pdct #2"),
    ("#3", "Top Pdct #3"),
    ("#4", "Top Pdct #4"),
    ("#5", "Top Pdct #5"),
]


def clean_domain(domain):
    return domain.strip().replace("https://", "").replace("http://", "").split("/")[0]


def normalize_date(value):
    return pd.to_datetime(value).date()


def shopify_datetime_bounds(date_from, date_to):
    start = pd.Timestamp(date_from).tz_localize(SHOP_TIMEZONE)
    end = (
        pd.Timestamp(date_to) + pd.Timedelta(hours=23, minutes=59, seconds=59)
    ).tz_localize(SHOP_TIMEZONE)
    return start.isoformat(), end.isoformat()


def parse_order_created_at(order):
    created_at = order.get("created_at")
    if not created_at:
        return None
    try:
        return pd.to_datetime(created_at, utc=True).tz_convert(SHOP_TIMEZONE)
    except Exception:
        return None


def get_next_link(link_header):
    if not link_header:
        return None
    for part in link_header.split(","):
        chunks = part.split(";")
        if len(chunks) < 2:
            continue
        url_part = chunks[0].strip()
        rel_part = ";".join(chunks[1:])
        if 'rel="next"' in rel_part:
            return url_part.strip("<> ")
    return None


def request_shopify_with_retry(session, url, headers, params=None, log=print):
    last_response = None
    for attempt in range(MAX_RETRIES + 1):
        try:
            response = session.get(url, headers=headers, params=params, timeout=30)
        except requests.RequestException as e:
            if attempt == MAX_RETRIES:
                raise e
            wait = min(2 ** attempt, 10)
            log(f"Connexion instable. Nouvelle tentative dans {wait:.1f}s...")
            time.sleep(wait)
            continue

        last_response = response

        if response.status_code == 429 and attempt < MAX_RETRIES:
            retry_after = response.headers.get("Retry-After")
            try:
                wait = float(retry_after) if retry_after else min(2 ** attempt + 1, 10)
            except ValueError:
                wait = min(2 ** attempt + 1, 10)
            wait = max(wait, REQUEST_DELAY_SECONDS)
            log(f"Rate limit 429. Pause de {wait:.1f}s...")
            time.sleep(wait)
            continue

        if 500 <= response.status_code < 600 and attempt < MAX_RETRIES:
            wait = min(2 ** attempt + 1, 10)
            log(f"Erreur Shopify {response.status_code}. Nouvelle tentative dans {wait:.1f}s...")
            time.sleep(wait)
            continue

        return response

    return last_response


def get_all_orders(domain, token, date_from, date_to, log=print):
    domain = clean_domain(domain)
    created_at_min, created_at_max = shopify_datetime_bounds(date_from, date_to)

    url = f"https://{domain}/admin/api/{API_VERSION}/orders.json"
    headers = {"X-Shopify-Access-Token": token.strip()}
    params = {
        "limit": 250,
        "status": "any",
        "created_at_min": created_at_min,
        "created_at_max": created_at_max,
    }

    orders = []
    session = requests.Session()

    while url:
        try:
            response = request_shopify_with_retry(
                session=session, url=url, headers=headers, params=params, log=log
            )
        except Exception as e:
            raise RuntimeError(f"Erreur connexion Shopify : {e}")

        if response.status_code == 200:
            data = response.json()
            orders.extend(data.get("orders", []))
            log(f"{len(orders)} commandes récupérées...")
            url = get_next_link(response.headers.get("Link", ""))
            params = None
            if url:
                time.sleep(REQUEST_DELAY_SECONDS)
        else:
            raise RuntimeError(f"Erreur API Shopify ({response.status_code}): {response.text[:300]}")

    return orders


def get_ga4_monthly_metrics(property_id, service_account_info_dict, date_from, date_to, log=print):
    if not property_id or not service_account_info_dict:
        return pd.DataFrame()

    try:
        from google.analytics.data_v1beta import BetaAnalyticsDataClient
        from google.analytics.data_v1beta.types import DateRange, Dimension, Metric, RunReportRequest
        from google.oauth2 import service_account
    except ImportError:
        log("Librairies GA4 manquantes (google-analytics-data google-auth)")
        return pd.DataFrame()

    try:
        credentials = service_account.Credentials.from_service_account_info(
            service_account_info_dict,
            scopes=["https://www.googleapis.com/auth/analytics.readonly"],
        )
        client = BetaAnalyticsDataClient(credentials=credentials)

        # --- Requête 1 : métriques globales par mois ---
        req_global = RunReportRequest(
            property=f"properties/{property_id.strip()}",
            dimensions=[Dimension(name="yearMonth")],
            metrics=[
                Metric(name="sessions"),
                Metric(name="totalUsers"),
                Metric(name="sessionsPerUser"),
                Metric(name="averageSessionDuration"),
                Metric(name="bounceRate"),
                Metric(name="addToCarts"),
                Metric(name="checkouts"),
            ],
            date_ranges=[DateRange(start_date=str(date_from), end_date=str(date_to))],
        )
        resp_global = client.run_report(req_global)

        # --- Requête 2 : sessions par mois ET par canal (defaultChannelGroup) ---
        req_channel = RunReportRequest(
            property=f"properties/{property_id.strip()}",
            dimensions=[
                Dimension(name="yearMonth"),
                Dimension(name="sessionDefaultChannelGroup"),
            ],
            metrics=[Metric(name="sessions")],
            date_ranges=[DateRange(start_date=str(date_from), end_date=str(date_to))],
        )
        resp_channel = client.run_report(req_channel)

    except Exception as e:
        log(f"GA4 non récupéré : {e}")
        return pd.DataFrame()

    # Mapping GA4 channel → nos libellés
    CHANNEL_MAP = {
        "organic search":   "Organic",
        "direct":           "Direct",
        "paid search":      "Paid",
        "paid social":      "Paid",
        "email":            "Email",
        "organic social":   "Social",
        "referral":         "Referral",
        "unassigned":       "Unassigned",
    }
    OUR_CHANNELS = ["Organic", "Direct", "Paid", "Email", "Social", "Referral", "Unassigned"]

    # Agrégation des sessions par canal
    channel_by_month = defaultdict(lambda: defaultdict(int))
    for row in resp_channel.rows:
        ym      = row.dimension_values[0].value
        channel = row.dimension_values[1].value.lower()
        sessions_val = int(float(row.metric_values[0].value or 0))
        mapped = CHANNEL_MAP.get(channel, "Unassigned")
        channel_by_month[f"{ym[:4]}-{ym[4:]}"][mapped] += sessions_val

    rows = []
    for row in resp_global.rows:
        year_month = row.dimension_values[0].value
        month_key  = f"{year_month[:4]}-{year_month[4:]}"
        values     = [metric.value for metric in row.metric_values]
        sessions        = float(values[0] or 0)
        duration_seconds = float(values[3] or 0)
        bounce_rate     = float(values[4] or 0)

        ch = channel_by_month.get(month_key, {})
        entry = {
            "Mois": month_key,
            "sessions":             round(sessions),
            "unique_visitors":      round(float(values[1] or 0)),
            "sessions_per_visitor": round(float(values[2] or 0), 2),
            "duration_seconds":     round(duration_seconds, 1),
            "duration_minutes":     round(duration_seconds / 60, 2),
            "bounce_pct":           round(bounce_rate * 100, 1),
            "add_to_cart":          round(float(values[5] or 0)),
            "checkout":             round(float(values[6] or 0)),
        }
        for c in OUR_CHANNELS:
            entry[f"src_{c}"] = ch.get(c, 0)
        rows.append(entry)

    return pd.DataFrame(rows)


def compute_monthly_metrics(orders, ga4_df=None, date_from=None, date_to=None):
    date_from = normalize_date(date_from) if date_from else None
    date_to = normalize_date(date_to) if date_to else None

    valid_orders = []
    for order in orders:
        if order.get("financial_status") in ("voided",):
            continue
        created_local = parse_order_created_at(order)
        if created_local is None:
            continue
        order_date = created_local.date()
        if date_from and order_date < date_from:
            continue
        if date_to and order_date > date_to:
            continue
        month_key = created_local.strftime("%Y-%m")
        valid_orders.append((order, month_key))

    monthly = defaultdict(lambda: {"orders": [], "customers": set(), "returning_customers": set()})
    customer_order_count = defaultdict(int)

    for order, _ in valid_orders:
        customer = order.get("customer") or {}
        customer_id = customer.get("id")
        if customer_id:
            customer_order_count[customer_id] += 1

    for order, month_key in valid_orders:
        monthly[month_key]["orders"].append(order)
        customer = order.get("customer") or {}
        customer_id = customer.get("id")
        if customer_id:
            monthly[month_key]["customers"].add(customer_id)
            if customer_order_count[customer_id] > 1:
                monthly[month_key]["returning_customers"].add(customer_id)

    rows = []
    for month_key in sorted(monthly.keys()):
        month_data = monthly[month_key]
        month_orders = month_data["orders"]
        nb_orders = len(month_orders)
        nb_customers = len(month_data["customers"])
        nb_returning = len(month_data["returning_customers"])
        new_customers = nb_customers - nb_returning

        gross_sales = sum(float(o.get("subtotal_price") or 0) for o in month_orders)
        discounts = sum(float(o.get("total_discounts") or 0) for o in month_orders)
        shipping = sum(
            float(o.get("total_shipping_price_set", {}).get("shop_money", {}).get("amount", 0) or 0)
            for o in month_orders
        )
        taxes = sum(float(o.get("total_tax") or 0) for o in month_orders)
        total_sales = sum(float(o.get("total_price") or 0) for o in month_orders)
        net_sales = gross_sales - discounts

        refund_amount = 0
        refund_count = 0
        for order in month_orders:
            for refund in order.get("refunds", []):
                for refund_item in refund.get("refund_line_items", []):
                    refund_amount += float(refund_item.get("subtotal") or 0)
                    refund_count += int(refund_item.get("quantity") or 0)

        products_sold = sum(
            int(line_item.get("quantity") or 0)
            for order in month_orders
            for line_item in order.get("line_items", [])
        )

        france_sales = 0
        export_sales = 0
        country_sales = defaultdict(float)
        for order in month_orders:
            address = order.get("shipping_address") or order.get("billing_address") or {}
            country = address.get("country_code", "").upper()
            amount = float(order.get("total_price") or 0)
            if country == "FR":
                france_sales += amount
            else:
                export_sales += amount
                if country:
                    country_sales[country] += amount

        top3_export = sorted(country_sales.items(), key=lambda x: x[1], reverse=True)[:3]

        product_qty = defaultdict(lambda: {"title": "", "qty": 0, "revenue": 0.0})
        for order in month_orders:
            for line_item in order.get("line_items", []):
                product_id = str(line_item.get("product_id", ""))
                quantity = int(line_item.get("quantity") or 0)
                price = float(line_item.get("price") or 0)
                product_qty[product_id]["title"] = line_item.get("title", product_id)
                product_qty[product_id]["qty"] += quantity
                product_qty[product_id]["revenue"] += price * quantity

        top5 = sorted(product_qty.values(), key=lambda x: x["revenue"], reverse=True)[:5]

        ratio_cde_client = round(nb_orders / nb_customers, 2) if nb_customers else 0
        ratio_pdct_cde = round(products_sold / nb_orders, 2) if nb_orders else 0
        ratio_pdct_client = round(products_sold / nb_customers, 2) if nb_customers else 0
        aov_ht = round(net_sales / nb_orders, 2) if nb_orders else 0
        aov_ttc = round(total_sales / nb_orders, 2) if nb_orders else 0
        ca_client_ht = round(net_sales / nb_customers, 2) if nb_customers else 0
        ca_client_ttc = round(total_sales / nb_customers, 2) if nb_customers else 0
        ca_new_clients = round((new_customers / nb_customers) * net_sales, 2) if nb_customers else 0
        pct_ca_new = round((new_customers / nb_customers) * 100, 1) if nb_customers else 0
        ltv = round(ca_client_ht * ratio_cde_client * 12, 2)
        refund_pct = round((refund_amount / gross_sales) * 100, 1) if gross_sales else 0
        discount_pct = round((discounts / gross_sales) * 100, 1) if gross_sales else 0

        rows.append({
            "Mois": month_key,
            "sessions": "", "unique_visitors": "", "sessions_per_visitor": "",
            "duration_seconds": "", "duration_minutes": "", "bounce_pct": "",
            "add_to_cart": "", "checkout": "", "conversion_pct": "",
            "# Clients": nb_customers,
            "# Clients récurrents": nb_returning,
            "# Nouveaux clients": new_customers,
            "# Commandes": nb_orders,
            "# Produits vendus": products_sold,
            "Ratio cdes/clients": ratio_cde_client,
            "Ratio produit/cde": ratio_pdct_cde,
            "Ratio produit/client": ratio_pdct_client,
            "Gross Sales (€)": round(gross_sales, 2),
            "Discounts (€)": round(discounts, 2),
            "Discounts %": discount_pct,
            "Net Sales (€)": round(net_sales, 2),
            "Shipping (€)": round(shipping, 2),
            "Taxes (€)": round(taxes, 2),
            "Total Sales (€)": round(total_sales, 2),
            "AOV HT (€)": aov_ht,
            "AOV TTC incl. ship (€)": aov_ttc,
            "CA/client HT (€)": ca_client_ht,
            "CA/client TTC (€)": ca_client_ttc,
            "CA Nvx clients HT (€)": ca_new_clients,
            "% CA Nvx clients": pct_ca_new,
            "Fréquence achat": ratio_cde_client,
            "LTV estimée (€)": ltv,
            "Retours (€)": round(refund_amount, 2),
            "Retours (#)": refund_count,
            "Retours %": refund_pct,
            "Net Sales après retours (€)": round(net_sales - refund_amount, 2),
            "France (€)": round(france_sales, 2),
            "Export (€)": round(export_sales, 2),
            "Export #1 pays": top3_export[0][0] if len(top3_export) > 0 else "",
            "Export #1 CA (€)": round(top3_export[0][1], 2) if len(top3_export) > 0 else 0,
            "Export #2 pays": top3_export[1][0] if len(top3_export) > 1 else "",
            "Export #2 CA (€)": round(top3_export[1][1], 2) if len(top3_export) > 1 else 0,
            "Export #3 pays": top3_export[2][0] if len(top3_export) > 2 else "",
            "Export #3 CA (€)": round(top3_export[2][1], 2) if len(top3_export) > 2 else 0,
            "Top Pdct #1": top5[0]["title"] if len(top5) > 0 else "",
            "Top Pdct #1 CA (€)": round(top5[0]["revenue"], 2) if len(top5) > 0 else 0,
            "Top Pdct #2": top5[1]["title"] if len(top5) > 1 else "",
            "Top Pdct #2 CA (€)": round(top5[1]["revenue"], 2) if len(top5) > 1 else 0,
            "Top Pdct #3": top5[2]["title"] if len(top5) > 2 else "",
            "Top Pdct #3 CA (€)": round(top5[2]["revenue"], 2) if len(top5) > 2 else 0,
            "Top Pdct #4": top5[3]["title"] if len(top5) > 3 else "",
            "Top Pdct #4 CA (€)": round(top5[3]["revenue"], 2) if len(top5) > 3 else 0,
            "Top Pdct #5": top5[4]["title"] if len(top5) > 4 else "",
            "Top Pdct #5 CA (€)": round(top5[4]["revenue"], 2) if len(top5) > 4 else 0,
        })

    df = pd.DataFrame(rows)

    if not df.empty:
        total_period_sales = df["Total Sales (€)"].sum()
        df["POIDS %"] = df["Total Sales (€)"].apply(
            lambda v: round((v / total_period_sales) * 100, 1) if total_period_sales else 0
        )
        df["% cdes Nx clients"] = df.apply(
            lambda row: round((row["# Nouveaux clients"] / row["# Commandes"]) * 100, 1)
            if row["# Commandes"] else 0, axis=1,
        )
        for i in range(1, 4):
            df[f"Export #{i}"] = df.apply(
                lambda row, i=i: f'{row[f"Export #{i} pays"]} - {row[f"Export #{i} CA (€)"]} €'
                if row[f"Export #{i} pays"] else "", axis=1,
            )

    if ga4_df is not None and not ga4_df.empty and not df.empty:
        df = df.merge(ga4_df, on="Mois", how="left", suffixes=("", "_ga4"))
        ga4_columns = [
            "sessions", "unique_visitors", "sessions_per_visitor",
            "duration_seconds", "duration_minutes", "bounce_pct",
            "add_to_cart", "checkout",
            "src_Organic", "src_Direct", "src_Paid",
            "src_Email", "src_Social", "src_Referral", "src_Unassigned",
        ]
        for col in ga4_columns:
            ga4_col = f"{col}_ga4"
            if ga4_col in df.columns:
                df[col] = df[ga4_col].fillna("")
                df = df.drop(columns=[ga4_col])
            elif col not in df.columns:
                df[col] = ""

        df["conversion_pct"] = df.apply(
            lambda row: round((row["# Commandes"] / row["sessions"]) * 100, 2)
            if isinstance(row["sessions"], (int, float)) and row["sessions"] else "",
            axis=1,
        )

    return df


def make_months_as_columns(df):
    if df.empty:
        return pd.DataFrame()
    df_by_month = df.set_index("Mois")
    months = list(df_by_month.index)
    rows = []
    for label, source_key in REPORT_ROWS:
        row = {"Indicateur": label}
        for month in months:
            if source_key and source_key in df_by_month.columns:
                row[month] = df_by_month.at[month, source_key]
            else:
                row[month] = ""
        rows.append(row)
    return pd.DataFrame(rows)


def build_excel(df):
    display_df = make_months_as_columns(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vue mensuelle"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    section_fill = PatternFill("solid", fgColor="D9E1F2")
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    euro_fmt = '#,##0.00 "€"'
    pct_fmt = '0.0"%"'
    num_fmt = '#,##0'
    ratio_fmt = '0.00'

    euro_labels = {
        "Gross sales", "Discounts", "Net Sales", "Shipping", "Taxes",
        "TOTAL SALES", "AOV HT", "AOV TTC (incl ship)", "CA / client HT",
        "CA / client TTC (incl ship)", "CA Nx clients HT", "LTV estimée",
        "Retours €", "NET SALES", "FRANCE", "EXPORT",
    }
    pct_labels = {
        "Bounce %", "Conversion %", "% cdes Nx clients", "Discounts %",
        "POIDS %", "% CA Nvx clients", "Retours %",
    }
    ratio_labels = {
        "Sessions / visitor", "Ratio cdes/clients", "Ratio produit / cde",
        "Ratio produit / client", "Frequence achat",
    }
    count_labels = {
        "Sessions", "Visteurs uniques", "Add to cart", "Checkout",
        "COMMANDES", "Commandes Nx clients", "# Clients uniques",
        "# Nx clients", "# Clients récurrents", "# Produits vendus", "Retours #",
    }
    section_labels = {"TOP 3 EXPORT", "TOP 5 PDCT"}

    headers = list(display_df.columns)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in display_df.iterrows():
        excel_row = row_idx + 2
        label = row["Indicateur"]

        for col_idx, header in enumerate(headers, start=1):
            value = row[header]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            if label == "":
                cell.border = Border()
                continue

            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "right", vertical="center"
            )

            if label in section_labels:
                cell.fill = section_fill
                cell.font = Font(name="Arial", bold=True, color="1F4E79", size=11)
                continue

            cell.font = bold_font if col_idx == 1 else normal_font

            if row_idx % 2 == 0:
                cell.fill = stripe_fill

            if col_idx > 1:
                if label in euro_labels:
                    cell.number_format = euro_fmt
                elif label in pct_labels:
                    cell.number_format = pct_fmt
                elif label in ratio_labels:
                    cell.number_format = ratio_fmt
                elif label in count_labels:
                    cell.number_format = num_fmt

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 34
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Rapport de comparaison : Mois M vs Mois M-1 an (même mois année précédente)
# ---------------------------------------------------------------------------

# (label affiché, clé dans le df, sous-label optionnel, type de valeur)
# type : "euro" | "pct" | "ratio" | "count" | "text" | "input" | None
COMPARISON_ROWS = [
    # (section_titre, label_col_A, sous_label_col_B, clé_df, type_valeur)
    ("#1 ACQUISITION", None, None, None, None),
    (None, "Visiteurs uniques",       None,         "unique_visitors",             "count"),
    (None, "Nouveaux visiteurs",      None,         None,                          None),
    (None, "Sources acquisition",     "Organic",    "src_Organic",                 "count"),
    (None, "",                        "Direct",     "src_Direct",                  "count"),
    (None, "",                        "Paid",       "src_Paid",                    "count"),
    (None, "",                        "Email",      "src_Email",                   "count"),
    (None, "",                        "Social",     "src_Social",                  "count"),
    (None, "",                        "Referral",   "src_Referral",                "count"),
    (None, "",                        "Unassigned", "src_Unassigned",              "count"),
    (None, "Duration min",            None,         "duration_minutes",            "ratio"),
    (None, "Bounce %",                None,         "bounce_pct",                  "pct"),
    ("_BLANK_", None, None, None, None),
    (None, "Budget Ads",              None,         "_budget_ads_",                "input"),
    (None, "ROAS",                    None,         "_roas_",                      "roas"),
    ("_BLANK_", None, None, None, None),

    ("#2 CONVERSION", None, None, None, None),
    (None, "Add to cart",             None,         "add_to_cart",                 "count"),
    (None, "Checkout",                None,         "checkout",                    "count"),
    (None, "Commandes",               None,         "# Commandes",                 "count"),
    (None, "Tx conversion",           None,         "conversion_pct",              "pct"),
    (None, "Commandes Nx clients",    None,         "# Nouveaux clients",          "count"),
    (None, "% cdes Nx clients",       None,         "% cdes Nx clients",           "pct"),
    ("_BLANK_", None, None, None, None),

    ("#3 PERFORMANCE CA", None, None, None, None),
    (None, "Gross sales",             None,         "Gross Sales (€)",             "euro"),
    (None, "Discounts",               None,         "Discounts (€)",               "euro"),
    (None, "Discounts %",             None,         "Discounts %",                 "pct"),
    ("_BLANK_", None, None, None, None),
    (None, "Net Sales",               None,         "Net Sales (€)",               "euro"),
    (None, "Shipping",                None,         "Shipping (€)",                "euro"),
    (None, "Taxes",                   None,         "Taxes (€)",                   "euro"),
    (None, "TOTAL SALES",             None,         "Total Sales (€)",             "euro"),
    ("_BLANK_", None, None, None, None),
    (None, "Panier moyen HT",         None,         "AOV HT (€)",                  "euro"),
    (None, "Panier moyen TTC + ship", None,         "AOV TTC incl. ship (€)",      "euro"),
    ("_BLANK_", None, None, None, None),
    (None, "Retours €",               None,         "Retours (€)",                 "euro"),
    (None, "Retours #",               None,         "Retours (#)",                 "count"),
    (None, "Retours %",               None,         "Retours %",                   "pct"),
    (None, "Net sales - retour",      None,         "Net Sales après retours (€)", "euro"),
    ("_BLANK_", None, None, None, None),

    ("#4 PERFORMANCE PRODUITS", None, None, None, None),
    (None, "# Produits vendus",       None,         "# Produits vendus",           "count"),
    (None, "Ratio cdes/clients",      None,         "Ratio cdes/clients",          "ratio"),
    (None, "Ratio produit / cde",     None,         "Ratio produit/cde",           "ratio"),
    ("_BLANK_", None, None, None, None),
    (None, "Top 10 produits Volume",  None,         None,                          None),
    ("_BLANK_", None, None, None, None),

    ("#5 GEO", None, None, None, None),
    (None, "Net Sales",               None,         "Net Sales (€)",               "euro"),
    (None, "Net Sales FR €",          None,         "France (€)",                  "euro"),
    (None, "Net Sales FR %",          None,         "_net_sales_fr_pct_",          "pct"),
    (None, "Net Sales Export €",      None,         "Export (€)",                  "euro"),
    (None, "Net Sales Export %",      None,         "_net_sales_export_pct_",      "pct"),
    ("_BLANK_", None, None, None, None),
    (None, "Top 3 CA pays hors France", None,       None,                          None),
    (None, "#1",                      None,         "Export #1",                   "text"),
    (None, "#2",                      None,         "Export #2",                   "text"),
    (None, "#3",                      None,         "Export #3",                   "text"),
]


def _vs_pct(m, m1):
    """Calcule le % de variation entre m (Mois M) et m1 (Mois M-1 an)."""
    try:
        m = float(m)
        m1 = float(m1)
        if m1 == 0:
            return None
        return round((m - m1) / abs(m1) * 100, 1)
    except (TypeError, ValueError):
        return None


def _get(row, key):
    """Retourne la valeur d'une clé dans un dict de métriques, ou None."""
    if row is None or key is None:
        return None
    return row.get(key)


def build_excel_comparison(full_df: "pd.DataFrame", month_m: str, months_to_compare=None) -> bytes:
    """
    Génère un Excel de comparaison cumulatif.

    Pour CHAQUE mois présent dans full_df (trié chronologiquement), on crée
    un bloc de 4 colonnes : [Mois M | Mois M-1 an | vs% | vide].
    Les colonnes fixes A et B (Indicateur / Sous-catégorie) restent à gauche.

    full_df  : DataFrame complet (une ligne par mois, incluant les mois de référence)
    month_m  : dernier mois ajouté (utilisé pour le titre de l'onglet)
    months_to_compare : liste optionnelle des mois à afficher en blocs de comparaison
    """
    MONTH_FR = ["", "Jan", "Fév", "Mar", "Avr", "Mai", "Jun",
                "Jul", "Aoû", "Sep", "Oct", "Nov", "Déc"]

    def month_label(m):
        y, mo = int(m[:4]), int(m[5:])
        return f"{MONTH_FR[mo]} {y}"

    def prev_year_month(m):
        y, mo = int(m[:4]), int(m[5:])
        return f"{y - 1}-{mo:02d}"

    # Liste triée des mois à afficher en blocs de comparaison
    all_months = sorted(months_to_compare or full_df["Mois"].tolist())
    df_idx     = full_df.set_index("Mois")

    def row_of(m):
        return df_idx.loc[m].to_dict() if m in df_idx.index else None

    # ---- Helpers ----
    def _clean(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    def geo_pct(row, num_key, denom_key):
        if row is None:
            return None
        try:
            num   = float(row.get(num_key)   or 0)
            denom = float(row.get(denom_key) or 0)
            return round(num / denom * 100, 1) if denom else None
        except Exception:
            return None

    def get_value(row, key):
        if key == "_net_sales_fr_pct_":
            return geo_pct(row, "France (€)", "Net Sales (€)")
        if key == "_net_sales_export_pct_":
            return geo_pct(row, "Export (€)", "Net Sales (€)")
        return _clean(_get(row, key))

    def vs_pct_calc(m, m1):
        try:
            m, m1 = float(m), float(m1)
            return round((m - m1) / abs(m1) * 100, 1) if m1 != 0 else None
        except (TypeError, ValueError):
            return None

    # ---- Workbook ----
    wb = Workbook()
    ws = wb.active
    ws.title = f"Comparaison {month_label(month_m)}"

    # Styles
    fw_bold   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    f_normal  = Font(name="Arial", size=10)
    f_bold    = Font(name="Arial", bold=True, size=10)
    f_italic  = Font(name="Arial", size=10, italic=True)
    f_section = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    f_pos     = Font(name="Arial", size=10, color="217346")
    f_neg     = Font(name="Arial", size=10, color="C00000")

    fill_header  = PatternFill("solid", fgColor="1F4E79")
    fill_section = PatternFill("solid", fgColor="F4B942")
    fill_stripe  = PatternFill("solid", fgColor="F2F2F2")
    fill_input   = PatternFill("solid", fgColor="FFFF99")

    thin   = Side(style="thin",   color="CCCCCC")
    medium = Side(style="medium", color="888888")
    b_norm = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
    b_sect = Border(left=medium, right=medium, top=medium, bottom=medium)

    euro_fmt  = '#,##0.00 "€"'
    pct_fmt   = '0.0"%"'
    ratio_fmt = '0.00'
    count_fmt = '#,##0'
    vs_fmt    = '+0.0"%";-0.0"%";0.0"%"'

    # ---- En-têtes ----
    # Col A = Indicateur (fixe), Col B = Sous-catégorie (fixe)
    # Puis pour chaque mois : [Mois M | Mois M-1 an | vs% | (vide)]
    FIXED_COLS = 2
    BLOCK_SIZE = 4   # Mois M | Mois M-1 an | vs% | vide

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    # Ligne 1 : titre de groupe sur les 3 colonnes de chaque bloc
    # Ligne 2 : sous-titres des colonnes
    for m_idx, m in enumerate(all_months):
        base_col = FIXED_COLS + 1 + m_idx * BLOCK_SIZE  # 1-indexed
        label_m  = month_label(m)
        label_m1 = month_label(prev_year_month(m))

        # Groupe : merge sur 3 colonnes (pas la colonne vide)
        ws.merge_cells(
            start_row=1, start_column=base_col,
            end_row=1,   end_column=base_col + 2
        )
        cell = ws.cell(row=1, column=base_col, value=f"{label_m} vs {label_m1}")
        cell.font      = fw_bold
        cell.fill      = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = b_norm

        # Sous-titres ligne 2
        for sub_col, sub_label in enumerate([label_m, label_m1, "vs %", ""], start=base_col):
            c = ws.cell(row=2, column=sub_col, value=sub_label)
            c.font      = fw_bold
            c.fill      = fill_header if sub_label else PatternFill()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if sub_label:
                c.border = b_norm
            ws.column_dimensions[get_column_letter(sub_col)].width = 14 if sub_label != "" else 2

    # Titres colonnes fixes lignes 1+2 (merged verticalement)
    for fix_col, fix_label in enumerate(["Indicateur", "Sous-catégorie"], start=1):
        ws.merge_cells(start_row=1, start_column=fix_col, end_row=2, end_column=fix_col)
        c = ws.cell(row=1, column=fix_col, value=fix_label)
        c.font      = fw_bold
        c.fill      = fill_header
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = b_norm

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.freeze_panes = "C3"

    # ---- Lignes de données (à partir de la ligne 3) ----
    excel_row = 3
    stripe    = False

    for entry in COMPARISON_ROWS:
        section, label, sublabel, df_key, val_type = entry

        # Ligne vide
        if section == "_BLANK_":
            excel_row += 1
            stripe = not stripe
            continue

        # Ligne de section
        if section and section.startswith("#"):
            total_cols = FIXED_COLS + len(all_months) * BLOCK_SIZE
            for c in range(1, total_cols + 1):
                cell = ws.cell(row=excel_row, column=c, value=section if c == 1 else "")
                cell.font      = f_section
                cell.fill      = fill_section
                cell.border    = b_sect
                cell.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[excel_row].height = 18
            excel_row += 1
            stripe = False
            continue

        # Colonnes fixes A et B
        is_input      = val_type == "input"
        is_roas       = df_key == "_roas_"
        is_subcategory = (label == "")

        col_a = label    if label    is not None else ""
        col_b = sublabel if sublabel is not None else ""

        for fix_col, fix_val in enumerate([col_a, col_b], start=1):
            cell = ws.cell(row=excel_row, column=fix_col, value=fix_val)
            cell.border    = b_norm
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if is_subcategory:
                cell.font = f_italic if fix_col == 2 else f_normal
            else:
                cell.font = f_bold if fix_col == 1 else f_normal
            fill_row = fill_input if is_input else (fill_stripe if stripe else None)
            if fill_row:
                cell.fill = fill_row

        # Colonnes de données pour chaque mois
        for m_idx, m in enumerate(all_months):
            base_col = FIXED_COLS + 1 + m_idx * BLOCK_SIZE
            m1       = prev_year_month(m)
            row_m    = row_of(m)
            row_m1   = row_of(m1)

            if is_input:
                val_m, val_m1 = 5000, 5000
            elif is_roas:
                budget = 5000
                ns_m  = _clean(_get(row_m,  "Net Sales (€)"))
                ns_m1 = _clean(_get(row_m1, "Net Sales (€)"))
                val_m  = round(float(ns_m)  / budget, 2) if ns_m  else None
                val_m1 = round(float(ns_m1) / budget, 2) if ns_m1 else None
            else:
                val_m  = get_value(row_m,  df_key) if df_key else None
                val_m1 = get_value(row_m1, df_key) if df_key else None

            vs = None
            if not is_input and val_type not in ("text", None) and df_key:
                vs = vs_pct_calc(val_m, val_m1)

            fill_row = fill_input if is_input else (fill_stripe if stripe else None)

            # Mois M
            c = ws.cell(row=excel_row, column=base_col, value=val_m)
            c.border    = b_norm
            c.font      = f_italic if is_subcategory else f_normal
            c.alignment = Alignment(horizontal="right", vertical="center")
            if fill_row: c.fill = fill_row
            _apply_num_fmt(c, val_type, is_roas, is_input, euro_fmt, pct_fmt, ratio_fmt, count_fmt)

            # Mois M-1 an
            c = ws.cell(row=excel_row, column=base_col + 1, value=val_m1)
            c.border    = b_norm
            c.font      = f_italic if is_subcategory else f_normal
            c.alignment = Alignment(horizontal="right", vertical="center")
            if fill_row: c.fill = fill_row
            _apply_num_fmt(c, val_type, is_roas, is_input, euro_fmt, pct_fmt, ratio_fmt, count_fmt)

            # vs %
            c = ws.cell(row=excel_row, column=base_col + 2, value=vs)
            c.border    = b_norm
            c.alignment = Alignment(horizontal="right", vertical="center")
            if vs is not None:
                c.number_format = vs_fmt
                c.font = f_pos if vs >= 0 else f_neg
            else:
                c.font = f_normal
            if fill_row: c.fill = fill_row

            # Colonne vide séparatrice
            ws.cell(row=excel_row, column=base_col + 3, value="")

        excel_row += 1
        stripe = not stripe

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _apply_num_fmt(cell, val_type, is_roas, is_input, euro_fmt, pct_fmt, ratio_fmt, count_fmt):
    if is_roas or is_input:
        cell.number_format = ratio_fmt
    elif val_type == "euro":
        cell.number_format = euro_fmt
    elif val_type == "pct":
        cell.number_format = pct_fmt
    elif val_type == "ratio":
        cell.number_format = ratio_fmt
    elif val_type == "count":
        cell.number_format = count_fmt
        
